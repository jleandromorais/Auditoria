import pdfplumber
import pandas as pd
import re
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill

# --- 1. FUNÇÃO DE LIMPEZA DE NÚMEROS ---
def limpar_valor_pdf(valor):
    """Transforma 'R$ 1.000,00' em 1000.00"""
    if not isinstance(valor, str): return valor
    limpo = re.sub(r'[^\d,\.-]', '', valor) # Tira letras e R$
    limpo = limpo.replace('.', '').replace(',', '.') # Formato US
    try:
        return float(limpo)
    except ValueError:
        return 0.0

# --- 2. SELEÇÃO DE ARQUIVOS (INTERFACE VISUAL) ---
def selecionar_arquivo(titulo, tipo):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    pasta_padrao = os.path.join(os.environ['USERPROFILE'], 'Downloads')
    
    caminho = filedialog.askopenfilename(
        initialdir=pasta_padrao,
        title=titulo,
        filetypes=tipo
    )
    return caminho

print("--- INICIANDO AUDITORIA ---")

# 1. Selecionar o PDF
print("1. Selecione o arquivo PDF...")
pdf_path = selecionar_arquivo("Selecione a NOTA TÉCNICA (PDF)", [("PDF Files", "*.pdf")])
if not pdf_path: exit()

# 2. Selecionar o Excel do Time
print("2. Selecione a Tabela Excel para conferir...")
excel_path = selecionar_arquivo("Selecione o EXCEL DO TIME", [("Excel Files", "*.xlsx")])
if not excel_path: exit()

# --- 3. EXTRAÇÃO DOS DADOS DO PDF (Focado no Quadro 4) ---
print("\n🔍 Lendo o PDF e extraindo dados oficiais...")
dados_pdf = []

with pdfplumber.open(pdf_path) as pdf:
    for pagina in pdf.pages:
        texto = pagina.extract_text()
        # Procura especificamente o Quadro 4 (Conta Gráfica)
        if "Quadro 4" in texto and "Apuração do Saldo" in texto:
            tabelas = pagina.extract_tables()
            if tabelas:
                df_pdf = pd.DataFrame(tabelas[0])
                
                # Limpeza básica da tabela extraída
                df_pdf.columns = df_pdf.iloc[0] # Cabeçalho
                df_pdf = df_pdf[1:]
                
                # Vamos pegar a coluna de Descrição (DADOS) e o TOTAL
                # Nota: Ajuste os nomes das colunas 'DADOS' e 'TOTAL' conforme o PDF lido
                try:
                    df_pdf['Item'] = df_pdf['DADOS'] 
                    df_pdf['Valor_PDF'] = df_pdf['TOTAL'].apply(limpar_valor_pdf)
                    
                    dados_pdf = df_pdf[['Item', 'Valor_PDF']]
                    print("✅ Quadro 4 encontrado e extraído com sucesso!")
                except KeyError:
                    print("⚠️ Aviso: As colunas do PDF não têm os nomes esperados. Verifique a leitura.")
                break

if len(dados_pdf) == 0:
    print("❌ Erro: Não encontrei o Quadro 4 no PDF.")
    exit()

# --- 4. LEITURA DO EXCEL DO TIME ---
print("🔍 Lendo o Excel do time...")
try:
    # Assume que o Excel do time tem colunas 'Item' e 'Valor_Esperado'
    # Se tiver outros nomes, mude aqui:
    df_excel_time = pd.read_excel(excel_path)
    
    # Padronizar nomes para garantir que cruza certo
    # (Adapte isto aos nomes reais das colunas do seu time)
    if 'Valor_Esperado' not in df_excel_time.columns:
         # Tenta adivinhar se a coluna se chama 'Valor', 'Total', etc.
         df_excel_time.rename(columns={'Valor': 'Valor_Esperado', 'Total': 'Valor_Esperado'}, inplace=True)

except Exception as e:
    print(f"❌ Erro ao abrir Excel: {e}")
    exit()

# --- 5. O CRUZAMENTO (A AUDITORIA) ---
print("\n⚔️ CRUZANDO DADOS (AUDITORIA)...")

# Junta as duas tabelas baseadas no nome do Item (ex: RPV, SCG)
auditoria = pd.merge(dados_pdf, df_excel_time, on='Item', how='inner')

# Calcula a diferença
# Vamos considerar uma margem de erro de 1 centavo (0.01) para arredondamentos
auditoria['Diferenca'] = auditoria['Valor_PDF'] - auditoria['Valor_Esperado']
auditoria['Status'] = auditoria['Diferenca'].apply(lambda x: 'ERRO ❌' if abs(x) > 0.05 else 'OK ✅')

# --- 6. GERAR RELATÓRIO FINAL ---
arquivo_saida = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'Relatorio_Final_Auditoria.xlsx')

# Formatação condicional para ficar bonito
with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    auditoria.to_excel(writer, index=False, sheet_name='Resultado')
    
    # Acessa a aba para pintar as células
    workbook = writer.book
    worksheet = writer.sheets['Resultado']
    
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Pinta as linhas
    for row in range(2, len(auditoria) + 2): # Começa na linha 2 (pula cabeçalho)
        cell_status = worksheet[f'E{row}'] # Coluna E é o Status
        
        if cell_status.value == 'ERRO ❌':
            cell_status.fill = vermelho
            worksheet[f'D{row}'].fill = vermelho # Pinta a diferença também
        else:
            cell_status.fill = verde

print("\n" + "="*50)
print(f"🚀 AUDITORIA CONCLUÍDA!")
print(f"📂 Relatório gerado em: {arquivo_saida}")
print("="*50)

os.startfile(arquivo_saida)