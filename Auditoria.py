import pdfplumber
import pandas as pd
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
<<<<<<< HEAD
=======
import numpy as np
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711

# --- CONFIGURAÇÕES ---
ANO_ALVO = "25"
MESES_ALVO = ["OUT", "NOV", "DEZ"]

# ==============================================================================
# 1. LIMPEZA E UTILITÁRIOS (AGORA LÊ A NFE 576)
# ==============================================================================

def limpar_numero_nf_bruto(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return ""
    texto = str(valor).upper().strip()
    if "-" in texto:
        texto = texto.split("-")[0]
    if "/" in texto:
        texto = texto.split("/")[0]
    nums = re.findall(r"\d+", texto.replace(".", ""))
    if nums:
        return str(int(nums[0]))
    return ""

def to_float(texto):
    """
<<<<<<< HEAD
    Converte string para float.
    - Aceita formatos BR: 1.234,56
    - Aceita formatos com OCR zoado: 10,695,379,17 -> 10695379.17
    """
    if pd.isna(texto) or texto == "":
        return 0.0
    if isinstance(texto, (int, float)):
        return float(texto)
    if not isinstance(texto, str):
        return 0.0

    clean = texto.replace(" ", "")

    # Evita falso positivo do NCM/SH (27112100) virando "valor"
    if "27112100" in clean.replace(".", ""):
        return 0.0

    # Corrige múltiplas vírgulas (separador de milhar errado do OCR)
    if clean.count(",") > 1:
        partes = clean.split(",")
        inteiro = "".join(partes[:-1])
        decimal = partes[-1]
        clean = f"{inteiro}.{decimal}"

    limpo = re.sub(r"[^\d.,]", "", clean)  # mantém apenas números, ponto e vírgula

    if "," in limpo and "." not in limpo:
        limpo = limpo.replace(",", ".")
    elif "," in limpo and "." in limpo:
        limpo = limpo.replace(".", "").replace(",", ".")

    try:
        val = float(limpo)
        # corta valores absurdos (lixo de OCR)
        if val > 5_000_000_000 or val in [2024.0, 2025.0, 2026.0]:
            return 0.0
=======
    Converte string para float, corrigindo erros bizarros de OCR da NFE 576.
    Ex: "10,695,379,17" -> vira 10695379.17
    """
    if pd.isna(texto) or texto == "": return 0.0
    if isinstance(texto, (int, float)): return float(texto)
    if not isinstance(texto, str): return 0.0
    
    clean = texto.replace(" ", "")
    if "27112100" in clean.replace(".", ""): return 0.0 
    
    # --- CORREÇÃO DE MÚLTIPLAS VÍRGULAS (CRUCIAL PARA NFE 576) ---
    # Se houver mais de uma vírgula, é garantido que é separador de milhar errado.
    if clean.count(',') > 1:
        # Separa tudo por vírgula
        partes = clean.split(',')
        # Junta a parte inteira (tudo menos o último pedaço)
        inteiro = "".join(partes[:-1])
        # Pega a parte decimal (último pedaço)
        decimal = partes[-1]
        # Remonta no formato correto para float (10695379.17)
        clean = f"{inteiro}.{decimal}"
    
    # Limpeza Padrão (para as outras notas normais)
    limpo = re.sub(r'[^\d.,]', '', clean) # Mantém apenas números, ponto e vírgula
    
    # Se ainda tiver vírgula (formato BR padrão 1000,00), troca por ponto
    if ',' in limpo and '.' not in limpo: 
        limpo = limpo.replace(',', '.')
    elif ',' in limpo and '.' in limpo: # Caso misto, remove ponto e troca vírgula
        limpo = limpo.replace('.', '').replace(',', '.')

    try:
        val = float(limpo)
        if val > 5000000000 or val in [2024.0, 2025.0, 2026.0]: return 0.0
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711
        return val
    except:
        return 0.0

def remover_area_transporte_agressivo(texto):
    texto_novo = re.sub(r"(TRANSPORTADOR.*?DADOS DO PRODUTO)", "", texto, flags=re.IGNORECASE | re.DOTALL)
    if len(texto_novo) == len(texto):
        texto_novo = re.sub(r"(TRANSPORTADOR.*?CÓDIGO)", "", texto, flags=re.IGNORECASE | re.DOTALL)
    return texto_novo

def make_unique_columns(cols):
    """Garante nomes únicos (muito comum ter 'ICMS' duas vezes no Excel)."""
    seen = {}
    out = []
    for c in cols:
        base = str(c).strip()
        if base in seen:
            seen[base] += 1
            out.append(f"{base}__{seen[base]}")
        else:
            seen[base] = 0
            out.append(base)
    return out

# ==============================================================================
<<<<<<< HEAD
# 2. EXTRAÇÃO DOS PDFs
=======
# 2. EXTRAÇÃO (COM LÓGICA DE RESGATE PARA NFE 576)
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711
# ==============================================================================

def extrair_dados_tanque_final(texto_bruto, nome_arquivo):
    info = {
<<<<<<< HEAD
        "Arquivo": nome_arquivo,
        "Tipo": "NF-e",
        "Nota": "",
        "Vol": 0.0,
        "Bruto": 0.0,
        "ICMS": 0.0,
        "PIS": 0.0,
        "COFINS": 0.0,
        "Liq_Calc": 0.0,
        "Pedagio": 0.0
    }

    # 1) tipo
    if any(x in texto_bruto.upper() for x in ["CONHECIMENTO DE TRANSPORTE", "DACTE", "CT-E", "TIPO DO CTE"]):
        info["Tipo"] = "CT-e"

    # 2) preparação
    texto_analise = texto_bruto
    if info["Tipo"] == "NF-e":
        texto_analise = remover_area_transporte_agressivo(texto_bruto)

    # 3) número da nota
    match_nf = re.search(r"(?:N[º°o\.]*|NUMERO|DOC\.|DOCUMENTO)\s*[:\.]?\s*(\d+(?:\.\d+)*)",
                         texto_analise, re.IGNORECASE)
    if match_nf:
        info["Nota"] = limpar_numero_nf_bruto(match_nf.group(1))

    if not info["Nota"]:
        match_chave = re.search(r"(\d{44})", texto_analise.replace(" ", ""))
=======
        'Arquivo': nome_arquivo, 'Tipo': 'NF-e', 'Nota': '',
        'Vol': 0.0, 'Bruto': 0.0, 'ICMS': 0.0, 'Liq_Calc': 0.0,
        
        'PIS': 0.0, 'COFINS': 0.0,
        
        'Val_Produtos': 0.0, 'Base_ICMS_ST': 0.0, 'Val_ICMS_ST': 0.0,
        'Val_IPI': 0.0, 'Frete_Nota': 0.0, 'Seguro_Nota': 0.0, 
        'Desconto_Nota': 0.0, 'Outras_Desp': 0.0,

        'Val_Receber': 0.0, 'Pedagio': 0.0
    }
    
    # 1. TIPO
    if any(x in texto_bruto.upper() for x in ["CONHECIMENTO DE TRANSPORTE", "DACTE", "CT-E", "TIPO DO CTE"]):
        info['Tipo'] = "CT-e"
    
    # 2. PREPARAÇÃO
    texto_analise = texto_bruto
    if info['Tipo'] == 'NF-e':
        texto_analise = remover_area_transporte_agressivo(texto_bruto)

    # 3. NOTA
    match_nf = re.search(r'(?:N[º°o\.]*|NUMERO|DOC\.|DOCUMENTO)\s*[:\.]?\s*(\d+(?:\.\d+)*)', texto_analise, re.IGNORECASE)
    if match_nf: 
        info['Nota'] = limpar_numero_nf_bruto(match_nf.group(1))
    
    if not info['Nota']:
        match_chave = re.search(r'(\d{44})', texto_analise.replace(" ", ""))
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711
        if match_chave:
            chave = match_chave.group(1)
            info["Nota"] = str(int(chave[25:34]))

<<<<<<< HEAD
    # ----------------------------
    # CT-e
    # ----------------------------
    if info["Tipo"] == "CT-e":
        termos_vol_cte = [
            r"PESO\s*REAL.*?([\d\.]+,\d{3,4})",
            r"PESO\s*CALC.*?([\d\.]+,\d{3,4})",
            r"PESO\s*BC.*?([\d\.]+,\d{3,4})",
            r"PESO\s*TAXADO.*?([\d\.]+,\d{3,4})",
            r"CARGA.*?([\d\.]+,\d{3,4})",
            r"CUBAGEM.*?([\d\.]+,\d{3,4})",
            r"QUANTIDADE.*?([\d\.]+,\d{3,4})",
            r"PESO\s*AFERIDO.*?([\d\.]+,\d{3,4})",
            r"([\d\.]+,\d{3,4})\s*KG",
            r"([\d\.]+,\d{3,4})\s*M3"
        ]
        for t in termos_vol_cte:
            m = re.search(t, texto_analise, re.IGNORECASE | re.DOTALL)
            if m:
                v = to_float(m.group(1))
                if v > 0:
                    info["Vol"] = v
                    break

        m_serv = re.search(r"(?:VALOR\s*TOTAL\s*(?:DO)?\s*SERVI[ÇC]O|TOTAL\s*DA\s*PRESTA[ÇC][ÃA]O).*?([\d\.]+,\d{2})",
                           texto_analise, re.IGNORECASE | re.DOTALL)
        if m_serv:
            info["Bruto"] = to_float(m_serv.group(1))

        m_ped = re.search(r"PED[ÁA]GIO.*?([\d\.]+,\d{2})", texto_analise, re.IGNORECASE | re.DOTALL)
        if m_ped:
            info["Pedagio"] = to_float(m_ped.group(1))

        # ICMS no CT-e também pode vir como I.C.M.S
        m_icms = re.search(r"VALOR\s*(?:DO)?\s*I\.?\s*C\.?\s*M\.?\s*S\.?(?!\s*ST).*?([\d\.]+,\d{2})",
                           texto_analise, re.IGNORECASE | re.DOTALL)
        if m_icms:
            info["ICMS"] = to_float(m_icms.group(1))

        # líquido CT-e: bruto - ICMS (se houver)
        if info["Bruto"] > 0:
            base = info["Bruto"]
            if 0 < info["ICMS"] < base:
                base -= info["ICMS"]
            info["Liq_Calc"] = base

        return info

    # ----------------------------
    # NF-e
    # ----------------------------

    # Volume (m³, litros etc.)
    padroes_nfe = [
        r"([\d\.]+,\d{1,4})\s*(?:M3|M³|NM3)\b",
        r"(?:M3|M³|NM3)\s*([\d\.]+,\d{1,4})\b",
        r"(?:VOL(?:UME)?|QUANTIDADE|QUANT|QTDE|QTD|QTD\.)\s*[:\.]?\s*([\d\.]+,\d{1,4})",
        r"([\d\.]+,\d{1,4})\s*(?:L|LT|LITROS)\b",
        r"PESO\s*L[IÍ]QUIDO.*?([\d\.]+,\d{1,4})"
    ]
    for pat in padroes_nfe:
        m = re.search(pat, texto_analise, re.IGNORECASE | re.DOTALL)
        if m:
            v = to_float(m.group(1))
            if v > 0:
                info["Vol"] = v
                break

    # Bruto / total
    candidatos = [
        r"VALOR\s*TOTAL\s*DA\s*NOTA.*?([\d\.,]{6,})",
        r"VALOR\s*TOTAL\s*DOS\s*PRODUTOS.*?([\d\.,]{6,})",
        r"VL\.\s*TOTAL\s*[:\.]?\s*([\d\.]+,\d{2})",
        r"VALOR\s*TOTAL\s*[:]\s*([\d\.]+,\d{2})",
    ]
    for rgx in candidatos:
        m = re.search(rgx, texto_analise, re.IGNORECASE | re.DOTALL)
        if m:
            info["Bruto"] = to_float(m.group(1))
            if info["Bruto"] > 0:
                break

    # fallback: maior valor monetário encontrado
    if info["Bruto"] == 0:
        todos_valores = re.findall(r"[\d\.]+,\d{2}", texto_analise)
        floats = sorted([to_float(v) for v in todos_valores if 0 < to_float(v) < 5_000_000_000], reverse=True)
        if floats:
            info["Bruto"] = floats[0]

    # ICMS (NF-e) — no DANFE normalmente vem numa linha com: "BASE ... ICMS ... ST ... Produtos"
    # Ex (linha seguinte): "155.603,24 18.672,39 0,00 0,00 155.603,24"
    info["ICMS"] = 0.0
    m_linha = re.search(r"BASE\s*DE\s*C[ÁA]LCULO\s*DO\s*I\.?\s*C\.?\s*M\.?\s*S\.?.*?\n([^\n]+)",
                        texto_analise, re.IGNORECASE | re.DOTALL)
    if not m_linha:
        m_linha = re.search(r"VALOR\s*DO\s*I\.?\s*C\.?\s*M\.?\s*S\.?.*?\n([^\n]+)",
                            texto_analise, re.IGNORECASE | re.DOTALL)
    if m_linha:
        nums = re.findall(r"[\d\.]+,\d{2}", m_linha.group(1))
        # normalmente: [BASE_ICMS, VALOR_ICMS, BASE_ST, VALOR_ST, TOTAL_PRODUTOS]
        if len(nums) >= 2:
            info["ICMS"] = to_float(nums[1])

    if info["ICMS"] == 0.0:
        # fallback: tenta achar a linha do "CÁLCULO DO IMPOSTO" e pegar o 2º número (base, ICMS, ...)
        m_calc = re.search(r"C[ÁA]LCULO\s*DO\s*IMPOSTO.*?\n([^\n]+)", texto_analise, re.IGNORECASE | re.DOTALL)
        if m_calc:
            linha = m_calc.group(1)
            nums = re.findall(r"[\d\.]+,\d{2}", linha)
            if len(nums) >= 2:
                info["ICMS"] = to_float(nums[1])

# PIS/COFINS (nem sempre aparece no DANFE; se não achar, fica 0)
    m_pis = re.search(r"VALOR\s*(?:DO)?\s*PIS.*?([\d\.,]+\d{2})", texto_analise, re.IGNORECASE | re.DOTALL)
    if m_pis:
        info["PIS"] = to_float(m_pis.group(1))

    m_cof = re.search(r"VALOR\s*(?:DA)?\s*COFINS.*?([\d\.,]+\d{2})", texto_analise, re.IGNORECASE | re.DOTALL)
    if m_cof:
        info["COFINS"] = to_float(m_cof.group(1))

    # Líquido (NF-e): bruto - ICMS - PIS - COFINS (quando existirem)
    if info["Bruto"] > 0:
        base = info["Bruto"]
        for t in ("ICMS", "PIS", "COFINS"):
            v = info.get(t, 0.0) or 0.0
            if 0 < v < base:
                base -= v
        info["Liq_Calc"] = base
=======
    # --- CT-E ---
    if info['Tipo'] == "CT-e":
        # Volume
        termos_vol_cte = [
            r'PESO\s*REAL.*?([\d\.]+,\d{3,4})',
            r'PESO\s*CALC.*?([\d\.]+,\d{3,4})',
            r'PESO\s*BC.*?([\d\.]+,\d{3,4})',
            r'PESO\s*TAXADO.*?([\d\.]+,\d{3,4})',
            r'CARGA.*?([\d\.]+,\d{3,4})',
            r'CUBAGEM.*?([\d\.]+,\d{3,4})',
            r'QUANTIDADE.*?([\d\.]+,\d{3,4})',
            r'PESO\s*AFERIDO.*?([\d\.]+,\d{3,4})',
            r'([\d\.]+,\d{3,4})\s*KG',
            r'([\d\.]+,\d{3,4})\s*M3'
        ]
        for t in termos_vol_cte:
            m = re.search(t, texto_analise, re.IGNORECASE | re.DOTALL)
            if m: 
                v = to_float(m.group(1))
                if v > 0: 
                    info['Vol'] = v
                    break
        
        # Financeiro
        m_serv = re.search(r'(?:VALOR\s*TOTAL\s*(?:DO)?\s*SERVI[ÇC]O|TOTAL\s*DA\s*PRESTA[ÇC][ÃA]O).*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
        if m_serv: info['Bruto'] = to_float(m_serv.group(1))
        
        m_rec = re.search(r'(?:VALOR\s*A\s*RECEBER|TOTAL\s*A\s*PAGAR).*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
        if m_rec: info['Val_Receber'] = to_float(m_rec.group(1))
        
        m_ped = re.search(r'PED[ÁA]GIO.*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
        if m_ped: info['Pedagio'] = to_float(m_ped.group(1))
        
        m_icms = re.search(r'VALOR\s*(?:DO)?\s*ICMS.*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
        if m_icms: info['ICMS'] = to_float(m_icms.group(1))

        if info['Val_Receber'] > 0:
            info['Liq_Calc'] = info['Val_Receber']
        elif info['Bruto'] > 0:
            base = info['Bruto']
            if info['ICMS'] > 0: base -= info['ICMS']
            info['Liq_Calc'] = base

    # --- NF-E (COM A CORREÇÃO DE RESGATE) ---
    else:
        # Volume
        padroes_nfe = [
            r'([\d\.]+,\d{1,4})\s*(?:M3|M³|NM3)',
            r'(?:M3|M³|NM3).*?([\d\.]+,\d{1,4})',
            r'(?:QUANTIDADE|QUANT|QTDE|QTD|QTD\.|OTDIE)\s*[:\.]?.*?([\d\.]+,\d{1,4})',
            r'([\d\.]+,\d{1,4})\s*(?:L|LT|LITROS)\b',
            r'PESO\s*L[IÍ]QUIDO.*?([\d\.]+,\d{1,4})'
        ]
        for pat in padroes_nfe:
            match = re.search(pat, texto_analise, re.IGNORECASE)
            if not match and '.*?' in pat:
                match = re.search(pat, texto_analise, re.IGNORECASE | re.DOTALL)
            if match:
                v = to_float(match.group(1))
                if v > 0:
                    info['Vol'] = v
                    break
        
        # Financeiro: Busca Padrão Primeiro (Segura)
        m_head = re.search(r'VALOR\s*TOTAL\s*[:].*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE)
        if m_head: info['Bruto'] = to_float(m_head.group(1))
        
        if info['Bruto'] == 0:
            m_vl = re.search(r'VL\.\s*TOTAL\s*[:\.]?\s*([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
            if m_vl: info['Bruto'] = to_float(m_vl.group(1))

        if info['Bruto'] == 0:
            # Tenta padrão normal 1.000,00
            m_std = re.search(r'VALOR\s*TOTAL\s*DA\s*NOTA.*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
            if m_std: info['Bruto'] = to_float(m_std.group(1))

        # --- AQUI É O PULO DO GATO PARA A NFE 576 ---
        # Se ainda for 0, tenta encontrar o padrão "Feio" (10,695,379,17)
        if info['Bruto'] == 0:
            # Procura por qualquer sequencia de numeros e virgulas/pontos grande
            m_ugly = re.search(r'VALOR\s*TOTAL\s*DA\s*NOTA.*?([\d\.,]{8,})', texto_analise, re.IGNORECASE | re.DOTALL)
            if m_ugly:
                # A função to_float atualizada vai resolver as virgulas
                info['Bruto'] = to_float(m_ugly.group(1))
            
            # Tenta também Total dos Produtos se Total da Nota falhar
            if info['Bruto'] == 0:
                m_prod_ugly = re.search(r'VALOR\s*TOTAL\s*DOS\s*PRODUTOS.*?([\d\.,]{8,})', texto_analise, re.IGNORECASE | re.DOTALL)
                if m_prod_ugly:
                    info['Bruto'] = to_float(m_prod_ugly.group(1))

        # Fallback Final (Maior Valor)
        if info['Bruto'] == 0:
            todos_valores = re.findall(r'[\d\.]+,\d{2}', texto_analise)
            floats = sorted([to_float(v) for v in todos_valores if to_float(v) < 5000000000], reverse=True)
            if floats:
                maior = floats[0]
                if info['Bruto'] < 100 and maior > info['Bruto']: info['Bruto'] = maior

        # ICMS e outros campos
        m_icms = re.search(r'VALOR\s*(?:DO)?\s*ICMS(?!.*ST).*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
        if m_icms: info['ICMS'] = to_float(m_icms.group(1))
        
        # Outros Campos (Regex mais flexível para pegar números feios também)
        campos_regex = {
            'Val_Produtos': r'VALOR\s*TOTAL\s*DOS\s*PRODUTOS.*?([\d\.,]+\d{2})',
            'Base_ICMS_ST': r'BASE\s*(?:DE)?\s*C[AÁ]LC(?:ULO)?\s*(?:DO)?\s*ICMS\s*ST.*?([\d\.,]+\d{2})',
            'Val_ICMS_ST':  r'VALOR\s*(?:DO)?\s*ICMS\s*ST.*?([\d\.,]+\d{2})',
            'Frete_Nota':   r'VALOR\s*(?:DO)?\s*FRETE.*?([\d\.,]+\d{2})',
            'Seguro_Nota':  r'VALOR\s*(?:DO)?\s*SEGURO.*?([\d\.,]+\d{2})',
            'Desconto_Nota':r'DESCONTO.*?([\d\.,]+\d{2})',
            'Outras_Desp':  r'OUTRAS\s*DESP.*?([\d\.,]+\d{2})',
            'Val_IPI':      r'VALOR\s*(?:DO)?\s*IPI.*?([\d\.,]+\d{2})'
        }
        for campo, regex in campos_regex.items():
            m = re.search(regex, texto_analise, re.IGNORECASE | re.DOTALL)
            if m: info[campo] = to_float(m.group(1))

        # Líquido NF-e
        if info['Bruto'] > 0:
            base = info['Bruto']
            if info['ICMS'] > 0 and info['ICMS'] < base: base -= info['ICMS']
            info['Liq_Calc'] = base

    # Comuns
    m_pis = re.search(r'VALOR\s*(?:DO)?\s*PIS.*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
    if m_pis: info['PIS'] = to_float(m_pis.group(1))
    
    m_cofins = re.search(r'VALOR\s*(?:DA)?\s*COFINS.*?([\d\.]+,\d{2})', texto_analise, re.IGNORECASE | re.DOTALL)
    if m_cofins: info['COFINS'] = to_float(m_cofins.group(1))
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711

    return info

# ==============================================================================
# 3. EXCEL (LÊ NOTAS + VOLUME + S/TRIBUTOS + TAXAS/VALORES)
# ==============================================================================

def carregar_excel(caminho):
    print(f"⏳ Lendo Excel ({MESES_ALVO})...")
    dados = []
    try:
        xls = pd.read_excel(caminho, sheet_name=None, header=None)

        for aba, df in xls.items():
            aba_upper = str(aba).upper()
            if ANO_ALVO not in aba_upper:
                continue
            if not any(mes in aba_upper for mes in MESES_ALVO):
                continue

            idx = -1
            for i, row in df.head(80).iterrows():
                linha = [str(x).upper() for x in row.values]
                if any("NOTA" in x for x in linha) and (any("S/TRIBUTOS" in x for x in linha) or any("C/TRIBUTOS" in x for x in linha) or any("TOTAL" in x for x in linha)):
                    idx = i
                    break

<<<<<<< HEAD
            if idx == -1:
                continue

            cols = make_unique_columns([str(c).upper().strip() for c in df.iloc[idx]])
            df.columns = cols
            df = df[idx + 1 :].copy()

            c_nf = next((c for c in df.columns if "NOTA" in c or c == "NF"), None)
            c_liq = next((c for c in df.columns if "S/TRIBUTOS" in c), None)
            c_vol = next((c for c in df.columns if "VOL" in c or "M³" in c or "M3" in c or "QTDE" in c or "QTD" in c or "QUANT" in c), None)

            # Para impostos pode existir duplicado (ICMS, ICMS__1). Pegamos o ÚLTIMO (normalmente o de "conforme XML").
            c_icms = [c for c in df.columns if c.startswith("ICMS")]
            c_pis  = [c for c in df.columns if c.startswith("PIS")]
            c_cof  = [c for c in df.columns if c.startswith("COFINS")]

            c_icms = c_icms[-1] if c_icms else None
            c_pis  = c_pis[-1]  if c_pis  else None
            c_cof  = c_cof[-1]  if c_cof  else None

            if not (c_nf and c_liq):
                continue

            temp = df.copy()
            temp["NF_Clean"] = temp[c_nf].apply(limpar_numero_nf_bruto)
            temp["Vol_Excel"] = temp[c_vol].apply(to_float) if c_vol else 0.0
            temp["Liq_Excel"] = temp[c_liq].apply(to_float)
            temp["ICMS_Excel"] = temp[c_icms].apply(to_float) if c_icms else 0.0
            temp["PIS_Excel"] = temp[c_pis].apply(to_float) if c_pis else 0.0
            temp["COFINS_Excel"] = temp[c_cof].apply(to_float) if c_cof else 0.0
            temp["Mes"] = aba

            temp = temp[temp["NF_Clean"] != ""]
            if not temp.empty:
                dados.append(temp[["NF_Clean", "Vol_Excel", "Liq_Excel", "ICMS_Excel", "PIS_Excel", "COFINS_Excel", "Mes"]])
                print(f"   ✅ Aba {aba}: Carregada")

        if not dados:
            return pd.DataFrame()
        return pd.concat(dados, ignore_index=True)

=======
                if c_nf and c_val:
                    temp = df.copy()
                    temp['NF_Clean'] = temp[c_nf].apply(limpar_numero_nf_bruto)
                    temp['Vol_Excel'] = temp[c_vol].apply(to_float) if c_vol else 0.0
                    temp['Liq_Excel'] = temp[c_val].apply(to_float)
                    temp['Mes'] = aba
                    
                    temp = temp[temp['NF_Clean'] != ""]
                    if not temp.empty:
                        dados.append(temp[['NF_Clean', 'Vol_Excel', 'Liq_Excel', 'Mes']])
                        print(f"   ✅ Aba {aba}: Carregada")
        
        if not dados: return pd.DataFrame()
        return pd.concat(dados)
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711
    except Exception as e:
        messagebox.showerror("Erro Excel", str(e))
        return pd.DataFrame()

# ==============================================================================
# 4. RELATÓRIO
# ==============================================================================

def gerar_relatorio(lista, saida=None):
    df = pd.DataFrame(lista)
    cols = [
<<<<<<< HEAD
        "Arquivo", "Tipo", "Mes", "Nota",
        "Vol PDF", "Vol Excel", "Diff Vol",
        "Bruto PDF", "ICMS PDF", "PIS", "COFINS",
        "ICMS Excel", "PIS Excel", "COFINS Excel",
        "Liq PDF (Calc)", "Liq Excel", "Diff R$", "Status"
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = "-"
    df = df[cols]

    ts = datetime.now().strftime("%H%M%S")
    if saida is None:
        saida = os.path.join(os.environ.get("USERPROFILE", os.getcwd()), "Downloads", f"Auditoria_Final_{ts}.xlsx")

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resultado")
        ws = writer.sheets["Resultado"]

        header_fill = PatternFill("solid", fgColor="203764")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        verde = PatternFill("solid", fgColor="C6EFCE")
        vermelho = PatternFill("solid", fgColor="FFC7CE")

        # Status está na última coluna
        status_col = cols.index("Status") + 1

        for row in ws.iter_rows(min_row=2):
            status = str(row[status_col - 1].value)
            cor = verde if "OK" in status else vermelho

            for cell in row:
                cell.fill = cor
                if isinstance(cell.value, (int, float)):
                    # colunas monetárias (a partir do Bruto PDF)
                    if cell.col_idx >= cols.index("Bruto PDF") + 1:
                        cell.number_format = 'R$ #,##0.00'
                    # volumes
                    if cell.col_idx in [cols.index("Vol PDF")+1, cols.index("Vol Excel")+1, cols.index("Diff Vol")+1]:
                        cell.number_format = '#,##0.000'

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

    print(f"\n✅ Relatório Salvo: {saida}")
    try:
        os.startfile(saida)
    except:
        pass

    return saida
=======
        'Arquivo', 'Tipo', 'Mes', 'Nota', 'Vol PDF', 'Vol Excel', 'Diff Vol', 
        'Bruto PDF', 'ICMS PDF', 'PIS', 'COFINS',
        'Val_Produtos', 'Base_ICMS_ST', 'Val_ICMS_ST', 'Val_IPI', 
        'Frete_Nota', 'Seguro_Nota', 'Desconto_Nota', 'Outras_Desp',
        'Val_Receber', 'Pedagio', 
        'Liq PDF (Calc)', 'Liq Excel', 'Diff R$', 'Status'
    ]
    
    for c in cols: 
        if c not in df.columns: df[c] = '-'
    df = df[cols]
    
    ts = datetime.now().strftime("%H%M%S")
    saida = os.path.join(os.environ['USERPROFILE'], 'Downloads', f'Auditoria_Final_Fix576_{ts}.xlsx')
    
    try:
        with pd.ExcelWriter(saida, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Resultado')
            ws = writer.sheets['Resultado']
            
            header_fill = PatternFill("solid", fgColor="203764")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
            
            verde = PatternFill("solid", fgColor="C6EFCE")
            vermelho = PatternFill("solid", fgColor="FFC7CE")
            
            for row in ws.iter_rows(min_row=2):
                status = str(row[24].value)
                cor = verde if "OK" in status else vermelho
                
                for cell in row:
                    cell.fill = cor
                    if isinstance(cell.value, (int, float)):
                        if cell.col_idx >= 8: cell.number_format = 'R$ #,##0.00'
                        if cell.col_idx in [5,6,7]: cell.number_format = '#,##0.000'
            
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 25

        print(f"\n✅ Relatório Salvo: {saida}")
        try: os.startfile(saida)
        except: pass
    except Exception as e: print(f"❌ Erro: {e}")
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711

# ==============================================================================
# MAIN
# ==============================================================================

<<<<<<< HEAD
def auditar(pdfs, excel_path, saida=None):
    df_base = carregar_excel(excel_path)
    if df_base.empty:
        raise RuntimeError("Não consegui carregar as abas do Excel alvo (verifique ANO_ALVO e MESES_ALVO).")
=======
def main():
    root = tk.Tk(); root.withdraw()
    print("--- AUDITORIA FINAL (FIX NFE 576) ---")
    
    pdfs = filedialog.askopenfilenames(title="1. PDFs", filetypes=[("PDF", "*.pdf")])
    if not pdfs: return
    
    excel = filedialog.askopenfilename(title="2. Excel", filetypes=[("Excel", "*.xlsx")])
    if not excel: return
    
    df_base = carregar_excel(excel)
    if df_base.empty: return
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711

    relatorio = []
    print("\n🔍 Analisando PDFs...")

    for pdf in pdfs:
        try:
            with pdfplumber.open(pdf) as p:
                texto = ""
                for page in p.pages:
                    t = page.extract_text() or ""
                    texto += t + "\n"
                info = extrair_dados_tanque_final(texto, os.path.basename(pdf))
        except Exception:
            continue

        item = info.copy()
        item["Vol Excel"] = 0.0
        item["Liq Excel"] = 0.0
        item["Diff Vol"] = 0.0
        item["Diff R$"] = 0.0
        item["Status"] = "Ñ ENCONTRADO ⚠️"
        item["Mes"] = "-"
        item["ICMS Excel"] = 0.0
        item["PIS Excel"] = 0.0
        item["COFINS Excel"] = 0.0

        if info["Nota"]:
            match = df_base[df_base["NF_Clean"] == info["Nota"]]
            if not match.empty:
                row = match.iloc[0]
<<<<<<< HEAD
                item["Vol Excel"] = float(row["Vol_Excel"])
                item["Liq Excel"] = float(row["Liq_Excel"])
                item["Mes"] = row["Mes"]

                item["ICMS Excel"] = float(row["ICMS_Excel"])
                item["PIS Excel"] = float(row["PIS_Excel"])
                item["COFINS Excel"] = float(row["COFINS_Excel"])

                # Ajuste do líquido:
                # - NF-e: usa ICMS/PIS/COFINS do PDF quando existir; se não existir (muito comum no DANFE),
                #         usa os valores do Excel (XML) para o cálculo ficar coerente com "S/TRIBUTOS".
                if info["Tipo"] == "NF-e" and info["Bruto"] > 0:
                    icms = info["ICMS"] if info["ICMS"] > 0 else item["ICMS Excel"]
                    pis = info["PIS"] if info["PIS"] > 0 else item["PIS Excel"]
                    cof = info["COFINS"] if info["COFINS"] > 0 else item["COFINS Excel"]
                    item["Liq_Calc"] = info["Bruto"] - icms - pis - cof
                    if item["Liq_Calc"] < 0:
                        item["Liq_Calc"] = 0.0

                v_pdf = info["Vol"]
                vol_excel_num = to_float(item["Vol Excel"])
                item["Diff Vol"] = v_pdf - vol_excel_num
                item["Diff R$"] = float(item["Liq_Calc"]) - float(item["Liq Excel"])

                tol_r = 50.0 if info["Tipo"] == "CT-e" else 5.0
                financeiro_ok = abs(item["Diff R$"]) < tol_r
                volume_ok = abs(item["Diff Vol"]) < 1.0
                excel_sem_vol = (vol_excel_num == 0)

                if financeiro_ok and (volume_ok or excel_sem_vol):
                    item["Status"] = "OK ✅"
                    if excel_sem_vol:
                        item["Vol Excel"] = "NÃO NO EXCEL"
                        item["Diff Vol"] = "-"
                else:
                    status = []
                    if not volume_ok and not excel_sem_vol:
                        status.append("VOL")
                    if not financeiro_ok:
                        status.append("VALOR")
                    if not status:
                        status.append("ERRO")
                    item["Status"] = f"ERRO {'+'.join(status)} ❌"

        item["Vol PDF"] = info["Vol"]
        item["Bruto PDF"] = info["Bruto"]
        item["ICMS PDF"] = info["ICMS"]
        item["Liq PDF (Calc)"] = item["Liq_Calc"]

=======
                item['Vol Excel'] = row['Vol_Excel']
                item['Liq Excel'] = row['Liq_Excel']
                item['Mes'] = row['Mes']
                
                v_pdf = info['Vol']
                vol_excel_raw = item['Vol Excel']
                vol_excel_num = to_float(vol_excel_raw) if not pd.isna(vol_excel_raw) else 0.0
                
                item['Diff Vol'] = v_pdf - vol_excel_num
                item['Diff R$'] = info['Liq_Calc'] - item['Liq Excel']
                
                tol_r = 50.0 if info['Tipo'] == 'CT-e' else 5.0
                financeiro_ok = abs(item['Diff R$']) < tol_r
                volume_ok = abs(item['Diff Vol']) < 1.0
                excel_sem_vol = (vol_excel_num == 0)

                if financeiro_ok and (volume_ok or excel_sem_vol):
                    item['Status'] = 'OK ✅'
                    if excel_sem_vol:
                        item['Vol Excel'] = "NÃO NO EXCEL"
                        item['Diff Vol'] = "-"
                else:
                    status = []
                    if not volume_ok and not excel_sem_vol: status.append("VOL")
                    if not financeiro_ok: status.append("VALOR")
                    if not status: status.append("ERRO GENÉRICO")
                    item['Status'] = f"ERRO {'+'.join(status)} ❌"
        
        item['Vol PDF'] = info['Vol']
        item['Bruto PDF'] = info['Bruto']
        item['ICMS PDF'] = info['ICMS']
        item['Liq PDF (Calc)'] = info['Liq_Calc']
        
>>>>>>> 266edd6cdc1c75a1629c224d0b9a97bc0873b711
        relatorio.append(item)
        print(f"-> {info['Tipo']} {info['Nota']}: {item['Status']}")

    return gerar_relatorio(relatorio, saida=saida)

def main():
    root = tk.Tk()
    root.withdraw()
    print("--- AUDITORIA FINAL (ICMS + LIQUIDO AJUSTADO) ---")

    pdfs = filedialog.askopenfilenames(title="1. PDFs", filetypes=[("PDF", "*.pdf")])
    if not pdfs:
        return

    excel = filedialog.askopenfilename(title="2. Excel", filetypes=[("Excel", "*.xlsx")])
    if not excel:
        return

    try:
        auditar(pdfs, excel)
    except Exception as e:
        messagebox.showerror("Erro", str(e))

if __name__ == "__main__":
    main()
